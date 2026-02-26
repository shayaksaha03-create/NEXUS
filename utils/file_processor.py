"""
NEXUS AI - File Processor
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Processes uploaded files (images, PDFs, videos, text, documents)
into formats the LLM can understand:

  • Images  → base64 for Ollama multimodal API
  • PDFs    → extracted text
  • Videos  → keyframes (base64) + metadata
  • Text    → raw content
  • Docs    → extracted text (best-effort)

Dependencies are auto-detected with graceful degradation.
"""

import base64
import mimetypes
import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional, Tuple
from enum import Enum

import sys
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from utils.logger import get_logger

logger = get_logger("file_processor")


# ═══════════════════════════════════════════════════════════════════════════════
# OPTIONAL DEPENDENCY DETECTION
# ═══════════════════════════════════════════════════════════════════════════════

HAS_PYMUPDF = False
try:
    import fitz  # PyMuPDF
    HAS_PYMUPDF = True
except ImportError:
    pass

HAS_OPENCV = False
try:
    import cv2
    HAS_OPENCV = True
except ImportError:
    pass

HAS_DOCX = False
try:
    import docx  # python-docx
    HAS_DOCX = True
except ImportError:
    pass

HAS_OPENPYXL = False
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    pass

HAS_PIL = False
try:
    from PIL import Image
    HAS_PIL = True
except ImportError:
    pass


# ═══════════════════════════════════════════════════════════════════════════════
# DATA TYPES
# ═══════════════════════════════════════════════════════════════════════════════

class FileType(Enum):
    IMAGE = "image"
    PDF = "pdf"
    VIDEO = "video"
    TEXT = "text"
    DOCUMENT = "document"     # docx, xlsx, pptx
    UNSUPPORTED = "unsupported"


@dataclass
class FileAttachment:
    """Processed file data ready for the LLM pipeline"""
    filepath: str                                # Original path
    filename: str                                # Basename
    file_type: FileType = FileType.UNSUPPORTED
    file_size: int = 0                           # Bytes
    mime_type: str = ""

    # ── Content depending on type ──
    extracted_text: str = ""                      # For PDFs, text, docs
    base64_images: List[str] = field(default_factory=list)  # For images/video frames
    image_descriptions: List[str] = field(default_factory=list)  # Metadata per image

    # ── Metadata ──
    page_count: int = 0                          # PDFs
    frame_count: int = 0                         # Videos
    duration_seconds: float = 0.0                # Videos
    width: int = 0                               # Images/video
    height: int = 0                              # Images/video
    error: str = ""                              # If processing failed
    success: bool = False

    @property
    def has_images(self) -> bool:
        return len(self.base64_images) > 0

    @property
    def has_text(self) -> bool:
        return len(self.extracted_text.strip()) > 0

    def get_context_text(self) -> str:
        """Get text context to inject into the LLM prompt"""
        parts = []
        if self.extracted_text:
            parts.append(
                f"[Attached File: {self.filename} "
                f"({self.file_type.value}, {self._human_size()})]\n"
                f"{self.extracted_text}"
            )
        elif self.has_images and not self.extracted_text:
            desc = f"[Attached Image: {self.filename} ({self._human_size()}"
            if self.width and self.height:
                desc += f", {self.width}x{self.height}px"
            desc += ")]"
            parts.append(desc)
        if self.error:
            parts.append(f"[File processing note: {self.error}]")
        return "\n".join(parts)

    def _human_size(self) -> str:
        size = self.file_size
        for unit in ("B", "KB", "MB", "GB"):
            if size < 1024:
                return f"{size:.1f}{unit}"
            size /= 1024
        return f"{size:.1f}TB"


# ═══════════════════════════════════════════════════════════════════════════════
# FILE TYPE DETECTION
# ═══════════════════════════════════════════════════════════════════════════════

IMAGE_EXTENSIONS = {
    ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".tiff", ".tif", ".ico"
}
VIDEO_EXTENSIONS = {
    ".mp4", ".avi", ".mkv", ".webm", ".mov", ".wmv", ".flv", ".m4v"
}
PDF_EXTENSIONS = {".pdf"}
TEXT_EXTENSIONS = {
    ".txt", ".md", ".py", ".js", ".ts", ".jsx", ".tsx", ".html", ".htm",
    ".css", ".json", ".xml", ".yaml", ".yml", ".toml", ".ini", ".cfg",
    ".csv", ".log", ".sh", ".bat", ".ps1", ".sql", ".r", ".go", ".rs",
    ".java", ".c", ".cpp", ".h", ".hpp", ".cs", ".rb", ".php", ".swift",
    ".kt", ".scala", ".lua", ".pl", ".ex", ".exs", ".hs", ".clj",
    ".dockerfile", ".makefile", ".gitignore", ".env",
}
DOCUMENT_EXTENSIONS = {".docx", ".xlsx", ".pptx"}

ALL_SUPPORTED = (
    IMAGE_EXTENSIONS | VIDEO_EXTENSIONS | PDF_EXTENSIONS |
    TEXT_EXTENSIONS | DOCUMENT_EXTENSIONS
)


def detect_file_type(filepath: str) -> FileType:
    """Detect the type of a file from its extension"""
    ext = Path(filepath).suffix.lower()
    if ext in IMAGE_EXTENSIONS:
        return FileType.IMAGE
    elif ext in PDF_EXTENSIONS:
        return FileType.PDF
    elif ext in VIDEO_EXTENSIONS:
        return FileType.VIDEO
    elif ext in TEXT_EXTENSIONS:
        return FileType.TEXT
    elif ext in DOCUMENT_EXTENSIONS:
        return FileType.DOCUMENT
    else:
        return FileType.UNSUPPORTED


def get_supported_extensions() -> list:
    """Return list of all supported file extensions"""
    return sorted(ALL_SUPPORTED)


def get_file_filter_string() -> str:
    """Get file filter string for QFileDialog"""
    img = " ".join(f"*{e}" for e in sorted(IMAGE_EXTENSIONS))
    vid = " ".join(f"*{e}" for e in sorted(VIDEO_EXTENSIONS))
    pdf = " ".join(f"*{e}" for e in sorted(PDF_EXTENSIONS))
    txt = " ".join(f"*{e}" for e in sorted(TEXT_EXTENSIONS))
    doc = " ".join(f"*{e}" for e in sorted(DOCUMENT_EXTENSIONS))
    all_ext = " ".join(f"*{e}" for e in sorted(ALL_SUPPORTED))
    return (
        f"All Supported Files ({all_ext});;"
        f"Images ({img});;"
        f"PDF Documents ({pdf});;"
        f"Videos ({vid});;"
        f"Text & Code ({txt});;"
        f"Office Documents ({doc});;"
        f"All Files (*)"
    )


# ═══════════════════════════════════════════════════════════════════════════════
# FILE PROCESSOR
# ═══════════════════════════════════════════════════════════════════════════════

class FileProcessor:
    """Processes files into LLM-ready content"""

    # Max sizes
    MAX_IMAGE_DIMENSION = 1024       # Resize images larger than this
    MAX_TEXT_CHARS = 50000           # Truncate text longer than this
    MAX_PDF_PAGES = 50              # Max PDF pages to extract
    MAX_VIDEO_FRAMES = 4            # Max keyframes to extract from video
    MAX_FILE_SIZE_MB = 100          # Skip files larger than this

    def process_file(self, filepath: str) -> FileAttachment:
        """
        Process a file and return a FileAttachment with extracted content.
        This is the main entry point.
        """
        path = Path(filepath)

        # ── Validate ──
        if not path.exists():
            return FileAttachment(
                filepath=filepath,
                filename=path.name,
                error=f"File not found: {filepath}",
                success=False
            )

        file_size = path.stat().st_size
        if file_size > self.MAX_FILE_SIZE_MB * 1024 * 1024:
            return FileAttachment(
                filepath=filepath,
                filename=path.name,
                file_size=file_size,
                error=f"File too large ({file_size / 1024 / 1024:.1f}MB, max {self.MAX_FILE_SIZE_MB}MB)",
                success=False
            )

        file_type = detect_file_type(filepath)
        mime, _ = mimetypes.guess_type(filepath)

        attachment = FileAttachment(
            filepath=str(path.absolute()),
            filename=path.name,
            file_type=file_type,
            file_size=file_size,
            mime_type=mime or "application/octet-stream"
        )

        try:
            if file_type == FileType.IMAGE:
                self._process_image(path, attachment)
            elif file_type == FileType.PDF:
                self._process_pdf(path, attachment)
            elif file_type == FileType.VIDEO:
                self._process_video(path, attachment)
            elif file_type == FileType.TEXT:
                self._process_text(path, attachment)
            elif file_type == FileType.DOCUMENT:
                self._process_document(path, attachment)
            else:
                attachment.error = (
                    f"Unsupported file type: {path.suffix}. "
                    f"Supported: {', '.join(sorted(ALL_SUPPORTED))}"
                )
                attachment.success = False
                return attachment

            attachment.success = True
            logger.info(
                f"Processed {file_type.value}: {path.name} "
                f"({attachment._human_size()})"
            )

        except Exception as e:
            attachment.error = f"Processing error: {str(e)}"
            attachment.success = False
            logger.error(f"File processing error for {path.name}: {e}")

        return attachment

    # ─────────────────────────────────────────────────────────────────────────
    # IMAGE PROCESSING
    # ─────────────────────────────────────────────────────────────────────────

    def _process_image(self, path: Path, attachment: FileAttachment):
        """Convert image to base64 for Ollama multimodal API"""
        if HAS_PIL:
            # Use PIL to resize if needed and get dimensions
            img = Image.open(path)
            attachment.width, attachment.height = img.size

            # Convert to RGB if necessary (e.g., RGBA PNGs)
            if img.mode in ("RGBA", "P", "LA"):
                img = img.convert("RGB")

            # Resize if too large
            if max(img.size) > self.MAX_IMAGE_DIMENSION:
                img.thumbnail(
                    (self.MAX_IMAGE_DIMENSION, self.MAX_IMAGE_DIMENSION),
                    Image.Resampling.LANCZOS
                )
                attachment.width, attachment.height = img.size

            # Convert to base64
            import io
            buffer = io.BytesIO()
            img.save(buffer, format="JPEG", quality=85)
            b64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
            attachment.base64_images.append(b64)
            attachment.image_descriptions.append(
                f"Image: {path.name} ({attachment.width}x{attachment.height})"
            )
        else:
            # Fallback: read raw bytes
            raw = path.read_bytes()
            b64 = base64.b64encode(raw).decode("utf-8")
            attachment.base64_images.append(b64)
            attachment.image_descriptions.append(f"Image: {path.name}")

        logger.info(f"Image processed: {path.name}")

    # ─────────────────────────────────────────────────────────────────────────
    # PDF PROCESSING
    # ─────────────────────────────────────────────────────────────────────────

    def _process_pdf(self, path: Path, attachment: FileAttachment):
        """Extract text from PDF"""
        if not HAS_PYMUPDF:
            attachment.error = (
                "PyMuPDF not installed. Install with: pip install PyMuPDF"
            )
            # Try raw read fallback
            attachment.extracted_text = (
                f"[PDF file: {path.name}, "
                f"{attachment.file_size} bytes. "
                f"Install PyMuPDF to extract text: pip install PyMuPDF]"
            )
            return

        doc = fitz.open(str(path))
        attachment.page_count = len(doc)

        text_parts = []
        pages_to_read = min(len(doc), self.MAX_PDF_PAGES)

        for page_num in range(pages_to_read):
            page = doc[page_num]
            page_text = page.get_text("text")
            if page_text.strip():
                text_parts.append(
                    f"--- Page {page_num + 1} ---\n{page_text}"
                )

        doc.close()

        full_text = "\n\n".join(text_parts)

        # Truncate if too long
        if len(full_text) > self.MAX_TEXT_CHARS:
            full_text = (
                full_text[:self.MAX_TEXT_CHARS] +
                f"\n\n[Truncated — {attachment.page_count} pages total, "
                f"showing first {self.MAX_TEXT_CHARS} characters]"
            )

        attachment.extracted_text = full_text
        logger.info(
            f"PDF processed: {path.name} ({attachment.page_count} pages, "
            f"{len(full_text)} chars)"
        )

    # ─────────────────────────────────────────────────────────────────────────
    # VIDEO PROCESSING
    # ─────────────────────────────────────────────────────────────────────────

    def _process_video(self, path: Path, attachment: FileAttachment):
        """Extract keyframes from video"""
        if not HAS_OPENCV:
            attachment.error = (
                "OpenCV not installed. Install with: pip install opencv-python"
            )
            attachment.extracted_text = (
                f"[Video file: {path.name}, {attachment._human_size()}. "
                f"Install OpenCV to extract frames: pip install opencv-python]"
            )
            return

        cap = cv2.VideoCapture(str(path))

        if not cap.isOpened():
            attachment.error = f"Could not open video: {path.name}"
            return

        # Get video info
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        fps = cap.get(cv2.CAP_PROP_FPS)
        attachment.width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        attachment.height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        attachment.duration_seconds = total_frames / fps if fps > 0 else 0

        # Calculate frame positions to extract (evenly spaced)
        num_frames = min(self.MAX_VIDEO_FRAMES, max(1, total_frames))
        if total_frames <= num_frames:
            frame_positions = list(range(total_frames))
        else:
            step = total_frames / num_frames
            frame_positions = [int(step * i) for i in range(num_frames)]

        for idx, frame_pos in enumerate(frame_positions):
            cap.set(cv2.CAP_PROP_POS_FRAMES, frame_pos)
            ret, frame = cap.read()
            if not ret:
                continue

            # Resize if too large
            h, w = frame.shape[:2]
            if max(h, w) > self.MAX_IMAGE_DIMENSION:
                scale = self.MAX_IMAGE_DIMENSION / max(h, w)
                frame = cv2.resize(
                    frame, (int(w * scale), int(h * scale)),
                    interpolation=cv2.INTER_AREA
                )

            # Encode to JPEG and base64
            _, buffer = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 80])
            b64 = base64.b64encode(buffer.tobytes()).decode("utf-8")
            attachment.base64_images.append(b64)

            timestamp = frame_pos / fps if fps > 0 else 0
            attachment.image_descriptions.append(
                f"Frame {idx + 1}/{num_frames} at "
                f"{timestamp:.1f}s"
            )
            attachment.frame_count += 1

        cap.release()

        # Add metadata as text context
        attachment.extracted_text = (
            f"[Video: {path.name} | "
            f"Duration: {attachment.duration_seconds:.1f}s | "
            f"Resolution: {attachment.width}x{attachment.height} | "
            f"Frames extracted: {attachment.frame_count}]"
        )

        logger.info(
            f"Video processed: {path.name} "
            f"({attachment.duration_seconds:.1f}s, "
            f"{attachment.frame_count} keyframes)"
        )

    # ─────────────────────────────────────────────────────────────────────────
    # TEXT PROCESSING
    # ─────────────────────────────────────────────────────────────────────────

    def _process_text(self, path: Path, attachment: FileAttachment):
        """Read text/code files"""
        # Try multiple encodings
        text = None
        for encoding in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
            try:
                text = path.read_text(encoding=encoding)
                break
            except (UnicodeDecodeError, UnicodeError):
                continue

        if text is None:
            attachment.error = "Could not decode text file"
            attachment.extracted_text = f"[Could not read: {path.name}]"
            return

        # Truncate
        if len(text) > self.MAX_TEXT_CHARS:
            text = (
                text[:self.MAX_TEXT_CHARS] +
                f"\n\n[Truncated — file is {len(text)} characters, "
                f"showing first {self.MAX_TEXT_CHARS}]"
            )

        # Add file type hint for code
        ext = path.suffix.lower().lstrip(".")
        lang_map = {
            "py": "Python", "js": "JavaScript", "ts": "TypeScript",
            "jsx": "React JSX", "tsx": "React TSX", "html": "HTML",
            "css": "CSS", "json": "JSON", "xml": "XML", "yaml": "YAML",
            "yml": "YAML", "md": "Markdown", "sql": "SQL", "sh": "Shell",
            "java": "Java", "c": "C", "cpp": "C++", "cs": "C#",
            "go": "Go", "rs": "Rust", "rb": "Ruby", "php": "PHP",
            "swift": "Swift", "kt": "Kotlin", "r": "R",
        }
        lang = lang_map.get(ext, "")
        lang_hint = f" ({lang})" if lang else ""

        attachment.extracted_text = (
            f"[File: {path.name}{lang_hint}]\n"
            f"```{ext}\n{text}\n```"
        )

        logger.info(f"Text file processed: {path.name} ({len(text)} chars)")

    # ─────────────────────────────────────────────────────────────────────────
    # DOCUMENT PROCESSING (DOCX, XLSX, PPTX)
    # ─────────────────────────────────────────────────────────────────────────

    def _process_document(self, path: Path, attachment: FileAttachment):
        """Extract text from office documents"""
        ext = path.suffix.lower()

        if ext == ".docx":
            self._process_docx(path, attachment)
        elif ext == ".xlsx":
            self._process_xlsx(path, attachment)
        elif ext == ".pptx":
            self._process_pptx(path, attachment)
        else:
            attachment.error = f"Unsupported document type: {ext}"

    def _process_docx(self, path: Path, attachment: FileAttachment):
        """Extract text from DOCX"""
        if not HAS_DOCX:
            attachment.error = (
                "python-docx not installed. Install with: pip install python-docx"
            )
            attachment.extracted_text = (
                f"[Word document: {path.name}. "
                f"Install python-docx to extract text: pip install python-docx]"
            )
            return

        doc = docx.Document(str(path))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        text = "\n\n".join(paragraphs)

        if len(text) > self.MAX_TEXT_CHARS:
            text = text[:self.MAX_TEXT_CHARS] + "\n\n[Truncated]"

        attachment.extracted_text = f"[Document: {path.name}]\n{text}"
        attachment.page_count = len(doc.sections)
        logger.info(
            f"DOCX processed: {path.name} ({len(paragraphs)} paragraphs)"
        )

    def _process_xlsx(self, path: Path, attachment: FileAttachment):
        """Extract text from XLSX"""
        if not HAS_OPENPYXL:
            attachment.error = (
                "openpyxl not installed. Install with: pip install openpyxl"
            )
            attachment.extracted_text = (
                f"[Spreadsheet: {path.name}. "
                f"Install openpyxl to extract data: pip install openpyxl]"
            )
            return

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        text_parts = []

        for sheet_name in wb.sheetnames[:10]:  # Max 10 sheets
            sheet = wb[sheet_name]
            text_parts.append(f"=== Sheet: {sheet_name} ===")

            rows = []
            for row in sheet.iter_rows(max_row=200, values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append(" | ".join(cells))

            text_parts.append("\n".join(rows))

        wb.close()

        text = "\n\n".join(text_parts)
        if len(text) > self.MAX_TEXT_CHARS:
            text = text[:self.MAX_TEXT_CHARS] + "\n\n[Truncated]"

        attachment.extracted_text = f"[Spreadsheet: {path.name}]\n{text}"
        logger.info(f"XLSX processed: {path.name}")

    def _process_pptx(self, path: Path, attachment: FileAttachment):
        """Best-effort text extraction from PPTX using zipfile"""
        try:
            import zipfile
            import xml.etree.ElementTree as ET

            text_parts = []
            with zipfile.ZipFile(str(path), 'r') as z:
                slide_files = sorted([
                    f for f in z.namelist()
                    if f.startswith("ppt/slides/slide") and f.endswith(".xml")
                ])

                for slide_file in slide_files[:50]:
                    with z.open(slide_file) as f:
                        tree = ET.parse(f)
                        root = tree.getroot()
                        # Extract all text from the XML
                        texts = []
                        for elem in root.iter():
                            if elem.text and elem.text.strip():
                                texts.append(elem.text.strip())
                        if texts:
                            slide_num = slide_file.split("slide")[-1].replace(".xml", "")
                            text_parts.append(
                                f"--- Slide {slide_num} ---\n" +
                                "\n".join(texts)
                            )

            text = "\n\n".join(text_parts)
            attachment.page_count = len(text_parts)

            if len(text) > self.MAX_TEXT_CHARS:
                text = text[:self.MAX_TEXT_CHARS] + "\n\n[Truncated]"

            attachment.extracted_text = (
                f"[Presentation: {path.name}, "
                f"{attachment.page_count} slides]\n{text}"
            )
            logger.info(
                f"PPTX processed: {path.name} "
                f"({attachment.page_count} slides)"
            )

        except Exception as e:
            attachment.error = f"Could not read PowerPoint: {e}"
            attachment.extracted_text = (
                f"[Presentation: {path.name}. Could not extract text: {e}]"
            )


# ═══════════════════════════════════════════════════════════════════════════════
# SINGLETON
# ═══════════════════════════════════════════════════════════════════════════════

file_processor = FileProcessor()


def get_dependency_status() -> dict:
    """Report which optional dependencies are available"""
    return {
        "PyMuPDF (PDFs)": HAS_PYMUPDF,
        "OpenCV (Videos)": HAS_OPENCV,
        "python-docx (Word)": HAS_DOCX,
        "openpyxl (Excel)": HAS_OPENPYXL,
        "Pillow (Image resize)": HAS_PIL,
    }
